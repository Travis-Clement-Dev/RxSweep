import csv
from pathlib import Path

from openpyxl import load_workbook

from rxsweep.pipeline import SweepResult
from rxsweep.exports import write_exports
from rxsweep.ingest import QuarantinedRow
from rxsweep.report import action_queue
from rxsweep.triage import Finding


def _finding(citation, severity="critical", source="recall", label="name_match", **kw):
    base = dict(
        item_name="Cefazolin Sodium",
        item_row=2,
        item_ndc="0409-4058-01",
        source=source,
        label=label,
        record={"recall_number": "D-0628-2025", "classification": "Class I"},
        severity=severity,
        severity_rationale="Class I recall: reasonable probability of serious harm",
        citation=citation,
    )
    base.update(kw)
    return Finding(**base)


def _result(tmp_path: Path) -> SweepResult:
    report = tmp_path / "report.html"
    report.write_text("<html>memo</html>")
    return SweepResult(
        run_id="testrun",
        run_dir=tmp_path,
        findings=[
            _finding(1),
            _finding(2, severity="high", source="shortage", label="exact_ndc",
                     item_name="Lorazepam", item_ndc="0641-6001-25",
                     record={"generic_name": "Lorazepam", "status": "Current"},
                     severity_rationale="Active shortage on a stocked item"),
        ],
        quarantined=[QuarantinedRow(row=38, reason="invalid ndc: 'BADNDC99'", raw={})],
        manual_review=[],
        unchecked=["shortages source unavailable: 39 items unchecked against shortages"],
        summary="Two findings require review [1][2].",
        meta=dict(
            csv_name="sample.csv", items_checked=39, run_ts="2026-07-09T23:00:00Z",
            months_back=24, model="claude-haiku-4-5", ai_available=True,
            audit_path=str(tmp_path / "audit.jsonl"),
        ),
        tiers={"critical": 1, "high": 1},
        report_path=report,
    )


def test_write_exports_produces_all_formats(tmp_path):
    paths = write_exports(_result(tmp_path), tmp_path)
    assert set(paths) == {"csv", "xlsx", "md", "report"}
    for p in paths.values():
        assert p.exists()


def test_csv_contents(tmp_path):
    paths = write_exports(_result(tmp_path), tmp_path)
    rows = list(csv.DictReader(open(paths["csv"])))
    assert len(rows) == 2
    assert rows[0]["item_name"] == "Cefazolin Sodium"
    assert rows[0]["severity"] == "critical"
    assert "enforcement.json" in rows[0]["fda_source_url"]


def test_xlsx_opens_with_frozen_header_and_meta_sheet(tmp_path):
    paths = write_exports(_result(tmp_path), tmp_path)
    wb = load_workbook(paths["xlsx"])
    ws = wb["Findings"]
    assert ws.freeze_panes == "A2"
    assert ws.max_row == 3  # header + 2 findings
    assert "Run" in wb.sheetnames


def test_markdown_is_ai_ready(tmp_path):
    paths = write_exports(_result(tmp_path), tmp_path)
    md = paths["md"].read_text()
    assert "### [1] Cefazolin Sodium (critical)" in md
    assert "FDA source record: https://api.fda.gov" in md
    assert "assume all results are unvalidated" in md
    assert "Unchecked (treat as unknown, not clear)" in md
    assert "pharmacist verifies" in md.lower() or "pharmacist" in md


def test_action_queue_verbs_uncapped():
    findings = [_finding(1)] + [
        _finding(i, severity="high", source="shortage", label="exact_ndc")
        for i in range(2, 12)
    ]
    queue = action_queue(findings)
    assert len(queue) == 11  # uncapped (contract v1.3, D11)
    assert queue[0]["text"].startswith("Verify lots for Cefazolin Sodium")
    assert queue[0]["tag"] == "Class I"
    assert queue[1]["text"].startswith("Confirm supply plan")


def test_memo_report_structure(tmp_path):
    from rxsweep.report import render_report

    r = _result(tmp_path)
    html = render_report(r.findings, r.quarantined, r.manual_review, r.unchecked,
                         r.summary, r.meta)
    assert "Memorandum" in html
    assert "Actions required" in html
    assert "Verified by:" in html
    assert "Exhibit A" in html
    assert "@media print" in html