"""Export suite: the artifacts pharmacy informaticists actually exchange.

CSV for the spreadsheet workflow, XLSX for circulation, Markdown for the
user's own AI assistant (citations intact), and the memo-format report for
committees (print to PDF from the browser). All are derived from the same
SweepResult, so every format tells the same story.
"""

import csv
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

from rxsweep.pipeline import SweepResult
from rxsweep.report import source_url
from rxsweep.triage import Finding

_COLUMNS = [
    "citation", "severity", "item_name", "item_ndc", "item_row",
    "source", "label", "severity_rationale", "ai_rationale", "fda_source_url",
]

_XLSX_TINTS = {
    "critical": "F6E4E1",
    "high": "F6E4E1",
    "moderate": "F5EBD8",
    "info": "E3F0E8",
}

DISCLAIMER = (
    'openFDA: "Do not rely on openFDA to make decisions regarding medical care. '
    'While we make every effort to ensure that data is accurate, you should '
    'assume all results are unvalidated."'
)


def _rows(findings: list[Finding]) -> list[dict]:
    return [
        {
            "citation": f.citation,
            "severity": f.severity,
            "item_name": f.item_name,
            "item_ndc": f.item_ndc or "",
            "item_row": f.item_row,
            "source": f.source,
            "label": f.label,
            "severity_rationale": f.severity_rationale,
            "ai_rationale": f.ai_rationale or "",
            "fda_source_url": source_url(f) or "",
        }
        for f in findings
    ]


def write_csv(result: SweepResult, path: Path) -> Path:
    with open(path, "w", newline="", encoding="utf-8") as fh:
        writer = csv.DictWriter(fh, fieldnames=_COLUMNS)
        writer.writeheader()
        writer.writerows(_rows(result.findings))
    return path


def write_xlsx(result: SweepResult, path: Path) -> Path:
    wb = Workbook()
    ws = wb.active
    ws.title = "Findings"
    ws.append([c.replace("_", " ").title() for c in _COLUMNS])
    for cell in ws[1]:
        cell.font = Font(bold=True)
    for row in _rows(result.findings):
        ws.append([row[c] for c in _COLUMNS])
        tint = _XLSX_TINTS.get(row["severity"])
        if tint:
            ws.cell(row=ws.max_row, column=2).fill = PatternFill(
                start_color=tint, end_color=tint, fill_type="solid"
            )
    widths = [9, 10, 28, 14, 9, 10, 12, 46, 40, 50]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = w
        for cell in ws[ws.cell(row=1, column=i).column_letter]:
            cell.alignment = Alignment(vertical="top", wrap_text=(w >= 40))
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    meta = wb.create_sheet("Run")
    for k, v in result.meta.items():
        meta.append([k, str(v)])
    meta.append(["disclaimer", DISCLAIMER])
    wb.save(path)
    return path


def write_markdown(result: SweepResult, path: Path) -> Path:
    m = result.meta
    lines = [
        "# RxSweep formulary sweep findings",
        "",
        f"- File: {m['csv_name']}",
        f"- Items checked: {m['items_checked']}",
        f"- Run: {m['run_ts']} (recall window {m['months_back']} months)",
        f"- AI triage: {m['model'] if m['ai_available'] else 'off'}",
        f"- Findings by severity: "
        + (", ".join(f"{t}={n}" for t, n in result.tiers.items()) or "none"),
        "",
        "Informational tool. A pharmacist verifies every finding before action. "
        "Not clinical advice.",
        "",
        "## Findings",
        "",
    ]
    for f in result.findings:
        url = source_url(f)
        lines += [
            f"### [{f.citation}] {f.item_name} ({f.severity})",
            f"- NDC: {f.item_ndc or 'n/a'} (formulary row {f.item_row})",
            f"- Source: {f.source} | match: {f.label}",
            f"- Basis: {f.severity_rationale}",
        ]
        if f.ai_rationale:
            lines.append(f"- AI match reasoning (verify): {f.ai_rationale}")
        if url:
            lines.append(f"- FDA source record: {url}")
        lines.append("")
    if result.manual_review:
        lines += ["## Needs manual review (not AI-adjudicated this run)", ""]
        lines += [
            f"- {c.item.name} (row {c.item.row}), {c.source}: {c.reason}"
            for c in result.manual_review
        ] + [""]
    if result.quarantined:
        lines += ["## Excluded rows", ""]
        lines += [f"- CSV line {q.row}: {q.reason}" for q in result.quarantined] + [""]
    if result.unchecked:
        lines += ["## Unchecked (treat as unknown, not clear)", ""]
        lines += [f"- {u}" for u in result.unchecked] + [""]
    lines += [
        "## Provenance",
        "",
        "Data: FDA recall enforcement reports, FDA drug shortages, FDA NDC "
        "directory via api.fda.gov. Every AI prompt and completion for this run "
        f"is logged verbatim in `{m['audit_path']}`.",
        "",
        f"> {DISCLAIMER}",
        "",
    ]
    path.write_text("\n".join(lines), encoding="utf-8")
    return path


def write_exports(result: SweepResult, run_dir: Path) -> dict[str, Path]:
    return {
        "csv": write_csv(result, run_dir / "findings.csv"),
        "xlsx": write_xlsx(result, run_dir / "findings.xlsx"),
        "md": write_markdown(result, run_dir / "findings.md"),
        "report": result.report_path,
    }
